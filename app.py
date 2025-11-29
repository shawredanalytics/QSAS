import streamlit as st
import base64
from pathlib import Path
import json
import os
import requests

ROOT = Path(__file__).parent

# GitHub Gist fallback helpers defined early to avoid NameError when used below
def _github_gist_id():
    try:
        return str(st.secrets.get("GITHUB_GIST_ID") or os.environ.get("GITHUB_GIST_ID") or "")
    except Exception:
        return ""

def _gh_headers():
    tok = str(st.secrets.get("GITHUB_TOKEN") or os.environ.get("GITHUB_TOKEN") or "")
    h = {"Accept": "application/vnd.github+json"}
    if tok:
        h["Authorization"] = f"Bearer {tok}"
    return h

def _github_get_gist_json(default=None):
    gid = _github_gist_id()
    if not gid:
        return default
    try:
        url = f"https://api.github.com/gists/{gid}"
        r = requests.get(url, headers=_gh_headers(), timeout=15)
        if r.status_code == 200:
            data = r.json()
            files = data.get("files") or {}
            file = files.get("grid_registrations.json") or next(iter(files.values()), None)
            if file and file.get("content"):
                return json.loads(file["content"])
        return default
    except Exception:
        return default

def _github_put_gist_json(payload: dict, filename: str = "grid_registrations.json"):
    gid = _github_gist_id()
    if not gid:
        return False, "no_gist"
    try:
        url = f"https://api.github.com/gists/{gid}"
        body = {"files": {filename: {"content": json.dumps(payload, ensure_ascii=False, indent=2)}}}
        r = requests.patch(url, headers=_gh_headers(), json=body, timeout=20)
        return (r.status_code in (200, 201)), r.status_code
    except Exception:
        return False, "exception"

def read_text(rel_path: str) -> str:
    p = ROOT / rel_path
    try:
        return p.read_text(encoding="utf-8")
    except Exception:
        return ""

def file_to_data_url(rel_path: str) -> str:
    p = ROOT / rel_path
    if not p.exists():
        return ""
    ext = p.suffix.lower()
    mime = {
        ".png": "image/png",
        ".svg": "image/svg+xml",
        ".jpg": "image/jpeg",
        ".jpeg": "image/jpeg",
    }.get(ext, "application/octet-stream")
    data = base64.b64encode(p.read_bytes()).decode("ascii")
    return f"data:{mime};base64,{data}"

def build_embedded_page(html_rel: str, bootstrap_js: str = ""):
    html = read_text(html_rel)
    css = read_text("assets/css/style.css")
    storage_js = read_text("assets/js/storage.js")
    branding_js = read_text("assets/js/branding.js")
    # Select the correct page script based on the HTML being embedded.
    # For index.html (Home), do not inject admin.js or user.js to avoid side effects.
    if "user.html" in html_rel:
        page_js = read_text("assets/js/user.js")
    elif "admin.html" in html_rel:
        page_js = read_text("assets/js/admin.js")
    elif "hq-admin.html" in html_rel:
        page_js = read_text("assets/js/hq-admin.js")
    else:
        page_js = ""
    # Assets mapping (include both encoded and unencoded key variants)
    assets_map = {
        "assets/QuXAT Logo Facebook.png": file_to_data_url("assets/QuXAT Logo Facebook.png"),
        "assets/QuXAT%20Logo%20Facebook.png": file_to_data_url("assets/QuXAT Logo Facebook.png"),
        "assets/QuXAT_Round_Seal.png": file_to_data_url("assets/QuXAT_Round_Seal.png"),
        "assets/Authorized Signatory.png": file_to_data_url("assets/Authorized Signatory.png"),
        "assets/Authorized%20Signatory.png": file_to_data_url("assets/Authorized Signatory.png"),
        "assets/img/quxat-logo.svg": file_to_data_url("assets/img/quxat-logo.svg"),
        "assets/Shawred Analytics Logo.png": file_to_data_url("assets/Shawred Analytics Logo.png"),
        "assets/Shawred%20Analytics%20Logo.png": file_to_data_url("assets/Shawred Analytics Logo.png"),
        "assets/Transparent Logo SAPLC.png": file_to_data_url("assets/Transparent Logo SAPLC.png"),
        "assets/Transparent%20Logo%20SAPLC.png": file_to_data_url("assets/Transparent Logo SAPLC.png"),
    }
    # Extract body content from original HTML
    lower = html.lower()
    start = lower.find("<body")
    if start != -1:
        start = lower.find(">", start) + 1
        end = lower.rfind("</body>")
        body = html[start:end] if end != -1 else html[start:]
    else:
        body = html
    # Inline asset references found in the body using data URLs so they render inside the embedded iframe
    try:
        for key, data_url in assets_map.items():
            if data_url:
                body = body.replace(key, data_url)
    except Exception:
        pass
    assets_js = f"<script>window.QSAS_ASSETS = {assets_map!r};</script>"
    # Compose embedded HTML
    composed = f"""
<!DOCTYPE html>
<html>
  <head>
    <meta charset=\"utf-8\" />
    <meta name=\"viewport\" content=\"width=device-width, initial-scale=1\" />
    <title>QuXAT Embedded</title>
    <style>{css}</style>
    {assets_js}
    <script>{storage_js}</script>
    <script>{branding_js}</script>
    {f"<script>{bootstrap_js}</script>" if bootstrap_js else ""}
  </head>
  <body>
    {body}
    <script src=\"https://cdnjs.cloudflare.com/ajax/libs/jspdf/2.5.1/jspdf.umd.min.js\"></script>
    <script>{page_js}</script>
  </body>
</html>
"""
    return composed

st.set_page_config(page_title="QuXAT Healthcare Organization Self Assessment", layout="wide", initial_sidebar_state="expanded")

# Hide Streamlit's default 3-dots app menu on the top-right
st.markdown(
    """
    <style>
    button[title="View app menu"] { display: none !important; visibility: hidden !important; }
    #MainMenu { display: none !important; visibility: hidden !important; }
    button[aria-label*="feedback"], button[title*="feedback"] { display: none !important; visibility: hidden !important; }
    a[href*="streamlit.app"], a[href*="streamlit.io"] { display: none !important; visibility: hidden !important; }
    /* Hide specific Streamlit toolbar actions, keep sidebar toggle visible */
    button[title*="Fork"], button[aria-label*="Fork"],
    button[title*="Deploy"], button[aria-label*="Deploy"],
    button[title*="Rerun"], button[aria-label*="Rerun"] {
      display: none !important; visibility: hidden !important;
    }
    </style>
    """,
    unsafe_allow_html=True,
)

# Query param helpers to support end-to-end navigation from embedded pages
def _get_query_params():
    try:
        # Streamlit >=1.34
        return dict(st.query_params)
    except Exception:
        try:
            return st.experimental_get_query_params()
        except Exception:
            return {}

def _sync_section_from_query():
    qp = _get_query_params()
    if not qp:
        return
    val = qp.get("section")
    if isinstance(val, list):
        val = val[0] if val else None
    if isinstance(val, str) and val:
        # Normalize URL-encoded variants: treat '+' as space for manual links
        try:
            normalized = val.replace('+', ' ').strip()
        except Exception:
            normalized = val
        st.session_state["section"] = normalized

def _set_query_section(value: str):
    try:
        try:
            st.query_params.clear()
        except Exception:
            pass
        st.query_params["section"] = value
    except Exception:
        try:
            st.experimental_set_query_params(section=value)
        except Exception:
            pass

def _no_persist_bootstrap_js() -> str:
    return """
    (function(){
      try {
        try { window.localStorage && window.localStorage.clear && window.localStorage.clear(); } catch(e){}
        var __mem = {};
        if (window.localStorage) {
          window.localStorage.setItem = function(k,v){ __mem[String(k)] = String(v); };
          window.localStorage.getItem = function(k){ return Object.prototype.hasOwnProperty.call(__mem,String(k)) ? __mem[String(k)] : null; };
          window.localStorage.removeItem = function(k){ delete __mem[String(k)]; };
          window.localStorage.clear = function(){ __mem = {}; };
        }
      } catch(e){}
    })();
    """

# Sidebar: render navigation buttons for Home and Admin
def render_sidebar_once():
    # Initialize section
    if "section" not in st.session_state:
        st.session_state["section"] = "Product based Quality Check"

    with st.sidebar:
        st.subheader("Navigation")
        try:
            st.caption(f"QSAS Portal v{globals().get('APP_VERSION', '3')}")
        except Exception:
            st.caption("QSAS Portal v3")
        # Primary navigation at the top (ordered)
        go_product = st.button("Quality of Life - Self Assessment", use_container_width=True)
        go_services = st.button("Quality based Self Check of Products & Services", use_container_width=True)
        go_home = st.button("QuXAT Organizational Self Assessment", use_container_width=True)
        go_advisory = st.button("QuXAT Advisory Services", use_container_width=True)

        # Visual separation, admin actions moved to the bottom area
        try:
            st.divider()
        except Exception:
            st.markdown("<hr>", unsafe_allow_html=True)

        st.subheader("Admin")
        go_admin = st.button("QSAS Admin Portal", use_container_width=True)
        try:
            st.divider()
        except Exception:
            st.markdown("<hr>", unsafe_allow_html=True)
        st.subheader("Reports")
        go_reports = st.button("QuXAT Reports", use_container_width=True)

    if go_product:
        st.session_state["section"] = "Product based Quality Check"
        _set_query_section("Product based Quality Check")
        st.rerun()
    if go_services:
        st.session_state["section"] = "Quality based Self Check of Products & Services"
        _set_query_section("Quality based Self Check of Products & Services")
        st.rerun()
    if go_home:
        st.session_state["section"] = "Home"
        _set_query_section("Home")
        st.rerun()
    if go_admin:
        st.session_state["section"] = "Admin"
        _set_query_section("Admin")
        st.rerun()
    if go_advisory:
        st.session_state["section"] = "QuXAT Advisory Services"
        _set_query_section("QuXAT Advisory Services")
        st.rerun()
    if go_reports:
        st.session_state["section"] = "QuXAT Reports"
        _set_query_section("QuXAT Reports")
        st.rerun()
    if go_product:
        st.session_state["section"] = "Product based Quality Check"
        _set_query_section("Product based Quality Check")
        st.rerun()

_sync_section_from_query()
render_sidebar_once()
section = st.session_state.get("section", "Home")
# Redirect removed pages to Home
if section in ("Healthcare Quality Grid", "Register for the Healthcare Quality Grid"):
    section = "Home"
mode = "Inline assets (Cloud)"
admin_username = st.session_state.get("admin_username", "")
admin_password = st.session_state.get("admin_password", "")
admin_auto_login = bool(st.session_state.get("admin_auto_login", False))

def build_home_html():
    css = read_text("assets/css/style.css")
    # Explicitly use the QuXAT Facebook PNG logo on the Home page
    logo_src = file_to_data_url("assets/QuXAT Logo Facebook.png")
    html = f"""
<!DOCTYPE html>
<html>
  <head>
    <meta charset=\"utf-8\" />
    <meta name=\"viewport\" content=\"width=device-width, initial-scale=1\" />
    <title>QuXAT Self Assessment</title>
    <style>{css}</style>
  </head>
  <body>
    <img class=\"brand-logo\" alt=\"QuXAT\" src=\"{logo_src}\" style=\"display:block; margin:0 auto; height:120px;\" />
  </body>
</html>
"""
    return html

def build_home_hero_html():
    css = read_text("assets/css/style.css")
    html = f"""
<!DOCTYPE html>
<html>
  <head>
    <meta charset=\"utf-8\" />
    <meta name=\"viewport\" content=\"width=device-width, initial-scale=1\" />
    <style>{css}</style>
  </head>
  <body>
    <div class=\"container\">
      <div class=\"card hero-card\"> 
        <h2>QuXAT Self‑Assessment Score (QSAS) — simple, credible, and actionable</h2>
        <p class=\"hint\">QSAS helps healthcare organizations in India self‑assess core quality practices, instantly view a Quality Self‑Assessment Score, and receive tailored guidance to improve. When ready, submit for Admin verification to receive a certified certificate.</p>
        <ul class=\"list\" style=\"margin-top:12px;\"> 
          <li><span class=\"item-title\">Quick</span><span class=\"item-sub\">Complete in minutes with clear, practical metrics.</span></li>
          <li><span class=\"item-title\">Actionable</span><span class=\"item-sub\">Immediate score and suggestions mapped to your maturity.</span></li>
          <li><span class=\"item-title\">Credible</span><span class=\"item-sub\">Optional Admin verification for a certified certificate.</span></li>
        </ul>
      </div>

      <div class=\"card\"> 
        <h3>Why QSAS for India?</h3>
        <ul class=\"list\" style=\"margin-top:8px;\"> 
          <li><span class=\"item-title\">Compliance readiness</span><span class=\"item-sub\">Aligns with national quality frameworks and prepares for external audits.</span></li>
          <li><span class=\"item-title\">Patient safety</span><span class=\"item-sub\">Surfaces risk areas and strengthens safety‑critical practices across care delivery.</span></li>
          <li><span class=\"item-title\">Continuous improvement</span><span class=\"item-sub\">Provides a repeatable score and guidance to track progress over time.</span></li>
        </ul>
      </div>

      <div class=\"card\"> 
        <h3>Get started in 3 simple steps</h3>
        <div class=\"steps-grid\">
          <div class=\"step-card step-1\">
            <div class=\"step-title\"><span class=\"step-chip\">📋</span>1) Conduct Self Assessment</div>
            <div class=\"step-sub\">Pick the most relevant quality checklist for your facility.</div>
          </div>
          <div class=\"step-card step-2\">
            <div class=\"step-title\"><span class=\"step-chip\">✅</span>2) Answer the Metrics</div>
            <div class=\"step-sub\">Select the practices you comply with — it only takes minutes.</div>
          </div>
          <div class=\"step-card step-3\">
            <div class=\"step-title\"><span class=\"step-chip\">📊</span>3) View Your QSAS Score</div>
            <div class=\"step-sub\">Download your report; optionally submit for Admin verification and certificate.</div>
          </div>
        </div>
        <p class=\"hint mt\"><strong>How the score is computed:</strong> QSAS uses up to 10 metrics per checklist. Each “Yes” contributes an equal share so the total score is out of 100 points.</p>
      </div>

      <div class=\"card\"> 
        <h3>QSAS Scoring Classifications — Healthcare Organizations in India</h3>
        <ul class=\"legend-list\"> 
          <li> 
            <div class=\"item-title\"><span class=\"badge badge-exemplary\">Exemplary</span> ≥ 90%</div>
            <div class=\"hint\">Well-established quality practices; maintain standardization and continuous improvement.</div>
          </li>
          <li> 
            <div class=\"item-title\"><span class=\"badge badge-strong\">Strong</span> ≥ 75%</div>
            <div class=\"hint\">Solid systems; target improvements through audits and PDSA cycles.</div>
          </li>
          <li> 
            <div class=\"item-title\"><span class=\"badge badge-developing\">Developing</span> ≥ 50%</div>
            <div class=\"hint\">Growing capabilities; formalize procedures, owners, and indicators.</div>
          </li>
          <li> 
            <div class=\"item-title\"><span class=\"badge badge-early\">Early</span> ≥ 25%</div>
            <div class=\"hint\">Foundational stage; establish governance, documentation, and regular audits.</div>
          </li>
          <li> 
            <div class=\"item-title\"><span class=\"badge badge-needs-improvement\">Needs Immediate Improvement</span> &lt; 25%</div>
            <div class=\"hint\">Critical gaps; address safety risks and define a 90-day remediation plan.</div>
          </li>
        </ul>
      </div>
    </div>

    <div class=\"site-footer\">© QuXAT — Quality Improvement Support</div>
  </body>
</html>
"""
    return html

if section == "Home":
    _set_query_section("Product based Quality Check")
    st.session_state["section"] = "Product based Quality Check"
    html_prod = build_embedded_page("product-quality.html")
    st.components.v1.html(html_prod, height=2600, scrolling=True)
elif section == "User Assessment":
    # Render the embedded User page at the very top (no extra Streamlit headers)
    # Pass through deep-link parameters (category/checklist) to the embedded page via localStorage
    qp = _get_query_params()
    raw_cat = qp.get("category")
    raw_chk = qp.get("checklist")
    cat = None
    chk = None
    if isinstance(raw_cat, list):
        cat = raw_cat[0] if raw_cat else None
    elif isinstance(raw_cat, str):
        cat = raw_cat
    if isinstance(raw_chk, list):
        chk = raw_chk[0] if raw_chk else None
    elif isinstance(raw_chk, str):
        chk = raw_chk
    js_bootstrap = """
    (function(){{
      try {{
        var cat = {cat};
        var chk = {chk};
        if (cat) localStorage.setItem('qsas_boot_category', String(cat));
        if (chk) localStorage.setItem('qsas_boot_checklist', String(chk));
      }} catch(e) {{}}
    }})();
    """.format(cat=repr(cat), chk=repr(chk))
    html_user = build_embedded_page("user.html", bootstrap_js=js_bootstrap)
    st.components.v1.html(html_user, height=2200, scrolling=True)
elif section == "Healthcare Quality Grid":
    # Page removed — redirect to Home
    _set_query_section("Home")
    st.session_state["section"] = "Home"
    html_index = build_embedded_page("index.html")
    st.components.v1.html(html_index, height=6000, scrolling=True)
elif section == "Register for the Healthcare Quality Grid":
    # Page removed — redirect to Home
    _set_query_section("Home")
    st.session_state["section"] = "Home"
    html_index = build_embedded_page("index.html")
    st.components.v1.html(html_index, height=6000, scrolling=True)
elif section == "QuXAT Advisory Services":
    html_adv = build_embedded_page("advisory.html")
    st.components.v1.html(html_adv, height=3800, scrolling=False)
elif section == "QuXAT Reports":
    html_reports = build_embedded_page("reports.html")
    st.components.v1.html(html_reports, height=3600, scrolling=False)
elif section in ("Product based Quality Check", "Quality of Life - Self Assessment"):
    html_prod = build_embedded_page("product-quality.html")
    st.components.v1.html(html_prod, height=2600, scrolling=True)
elif section == "Quality based Self Check of Products & Services":
    html_srv = build_embedded_page("service-quality.html")
    st.components.v1.html(html_srv, height=2600, scrolling=True)
elif section == "Gap Assessment":
    qp = _get_query_params()
    raw_plan = qp.get("plan")
    plan = None
    if isinstance(raw_plan, list):
        plan = raw_plan[0] if raw_plan else None
    elif isinstance(raw_plan, str):
        plan = raw_plan
    js_bootstrap = """
    (function(){
      try {
        // Clear any previous plan to avoid stale reads
        try { localStorage.removeItem('qsas_gap_plan'); } catch(e) {}
        var plan = %s;
        // Expose boot plan for iframe scripts as a direct variable to avoid storage timing issues
        try { window.QSAS_BOOT_PLAN = plan; } catch(e) {}
        if (plan) localStorage.setItem('qsas_gap_plan', String(plan));
      } catch(e) {}
    })();
    """ % (repr(plan))
    html_gap = build_embedded_page("gap-assessment.html", bootstrap_js=js_bootstrap)
    st.components.v1.html(html_gap, height=3800, scrolling=False)
elif section == "Certificate":
    _set_query_section("Home")
    st.session_state["section"] = "Home"
    html_index = build_embedded_page("index.html")
    st.components.v1.html(html_index, height=4200, scrolling=False)
else:  # Admin
    # Render the embedded Admin page at the very top (no extra Streamlit headers)
    if mode == "Local iframe":
        st.components.v1.html(
            '<iframe src="http://localhost:8000/admin.html" style="width:100%; height:100vh; border:none;"></iframe>',
            height=1800,
            scrolling=False,
        )
    else:
        # Inject admin credentials and optional auto-login into embedded Admin page
        js_bootstrap = """
        (function(){{
          try {{
            const K={{u:'qsas_portal_username',p:'qsas_portal_password'}};
            const u={u};
            const p={p};
            if (u) localStorage.setItem(K.u, u);
            if (p) localStorage.setItem(K.p, p);
          }} catch(e) {{}}
          if ({auto_login}) {{
            window.addEventListener('load', function(){{
              try {{
                const form = document.getElementById('loginForm');
                const uEl = document.getElementById('adminUsername');
                const pEl = document.getElementById('adminPassword');
                if (uEl) uEl.value = {u};
                if (pEl) pEl.value = {p};
                if (form) form.dispatchEvent(new Event('submit', {{ bubbles: true, cancelable: true }}));
              }} catch(e) {{}}
            }});
          }}
        }})();
        """.format(u=repr(admin_username), p=repr(admin_password), auto_login=str(bool(admin_auto_login)).lower())
        # Handle GitHub sync via query params
        qp2 = _get_query_params()
        sync = qp2.get("sync")
        payload_b64 = qp2.get("payload")
        if isinstance(sync, list):
            sync = sync[0] if sync else None
        if isinstance(payload_b64, list):
            payload_b64 = payload_b64[0] if payload_b64 else None
        sync_msg = ""
        if sync == "grid" and payload_b64:
            try:
                data_json = json.loads(base64.b64decode(payload_b64).decode("utf-8"))
                ok = _github_put_json("data/grid_registrations.json", data_json, message="QSAS: sync grid registrations")
                if not ok:
                    ok2, code2 = _github_put_gist_json(data_json, filename="grid_registrations.json")
                    sync_msg = "ok" if ok2 else f"error:{code2}"
                else:
                    try:
                        _github_put_json("data/qsas_grid_registrations_backup.json", data_json, message="QSAS: update registrations backup")
                    except Exception:
                        pass
                    sync_msg = "ok"
            except Exception:
                sync_msg = "error"
        elif sync == "certs" and payload_b64:
            try:
                data_json = json.loads(base64.b64decode(payload_b64).decode("utf-8"))
                ok = _github_put_json("data/cert_issuances.json", data_json, message="QSAS: sync cert issuances")
                if not ok:
                    ok2, code2 = _github_put_gist_json(data_json, filename="cert_issuances.json")
                    sync_msg = "ok" if ok2 else f"error:{code2}"
                else:
                    sync_msg = "ok"
            except Exception:
                sync_msg = "error"
        # Bootstrap approved registrations from GitHub to localStorage for consistent cross-device data
        gh_regs = []
        try:
            gh_regs = _github_get_json("data/grid_registrations.json", default=[])
            if not gh_regs:
                gh_regs = _github_get_gist_json(default=[])
        except Exception:
            gh_regs = _github_get_gist_json(default=[])
        gh_boot = f"(function(){{try{{localStorage.setItem('qsas_grid_registrations'," + json.dumps(json.dumps(gh_regs)) + ");" + (f"localStorage.setItem('qsas_sync_result','{sync_msg}');" if sync_msg else "") + "}}catch(e){{}}}})();"
        html_admin = build_embedded_page("admin.html", bootstrap_js=js_bootstrap + "\n" + gh_boot)
        st.components.v1.html(html_admin, height=2200, scrolling=True)
def _github_repo():
    return str(st.secrets.get("GITHUB_REPO") or "shawredanalytics/QSAS")

def _github_token():
    return str(st.secrets.get("GITHUB_TOKEN") or os.environ.get("GITHUB_TOKEN") or "")

def _gh_headers():
    tok = _github_token()
    h = {"Accept": "application/vnd.github+json"}
    if tok:
        h["Authorization"] = f"Bearer {tok}"
    return h

def _github_get_json(path: str, default=None):
    try:
        owner_repo = _github_repo()
        url = f"https://api.github.com/repos/{owner_repo}/contents/{path}"
        r = requests.get(url, headers=_gh_headers(), timeout=15)
        if r.status_code == 200:
            data = r.json()
            content_b64 = data.get("content") or ""
            content = base64.b64decode(content_b64).decode("utf-8")
            return json.loads(content)
        return default
    except Exception:
        return default

def _github_default_branch():
    try:
        owner_repo = _github_repo()
        url = f"https://api.github.com/repos/{owner_repo}"
        r = requests.get(url, headers=_gh_headers(), timeout=15)
        if r.status_code == 200:
            return r.json().get("default_branch") or "main"
    except Exception:
        pass
    return "main"

def _github_gist_id():
    try:
        return str(st.secrets.get("GITHUB_GIST_ID") or os.environ.get("GITHUB_GIST_ID") or "")
    except Exception:
        return ""

def _github_get_gist_json(default=None):
    gid = _github_gist_id()
    if not gid:
        return default
    try:
        url = f"https://api.github.com/gists/{gid}"
        r = requests.get(url, headers=_gh_headers(), timeout=15)
        if r.status_code == 200:
            data = r.json()
            files = data.get("files") or {}
            # Prefer a file named grid_registrations.json
            file = files.get("grid_registrations.json") or next(iter(files.values()), None)
            if file and file.get("content"):
                return json.loads(file["content"])
        return default
    except Exception:
        return default

def _github_put_gist_json(payload: dict):
    gid = _github_gist_id()
    if not gid:
        return False, "no_gist"
    try:
        url = f"https://api.github.com/gists/{gid}"
        body = {
            "files": {
                "grid_registrations.json": {"content": json.dumps(payload, ensure_ascii=False, indent=2)}
            }
        }
        r = requests.patch(url, headers=_gh_headers(), json=body, timeout=20)
        return (r.status_code in (200, 201)), r.status_code
    except Exception:
        return False, "exception"

def _github_put_json(path: str, payload: dict, message: str = "QSAS: sync grid registrations"):
    owner_repo = _github_repo()
    url = f"https://api.github.com/repos/{owner_repo}/contents/{path}"
    # Get SHA if file exists
    sha = None
    try:
        r0 = requests.get(url, headers=_gh_headers(), timeout=15)
        if r0.status_code == 200:
            sha = r0.json().get("sha")
    except Exception:
        pass
    content = json.dumps(payload, ensure_ascii=False, indent=2)
    body = {
        "message": message,
        "content": base64.b64encode(content.encode("utf-8")).decode("ascii"),
        "branch": _github_default_branch(),
    }
    if sha:
        body["sha"] = sha
    r = requests.put(url, headers=_gh_headers(), json=body, timeout=20)
    return r.status_code in (200, 201)
 
def _load_nabl_labs():
    try:
        xlsx_path = ROOT / "data" / "NABL Accredited Labs 2025.xlsx"
        if not xlsx_path.exists():
            return []
        # Try pandas first
        try:
            import pandas as pd  # type: ignore
            df = pd.read_excel(str(xlsx_path))
            records = []
            for _, row in df.iterrows():
                def pick(*names):
                    for n in names:
                        if n in df.columns and pd.notna(row[n]):
                            return str(row[n])
                    return ""
                name = pick("Laboratory Name", "Lab Name", "Organization", "Name")
                state = pick("State", "STATE")
                district = pick("District", "DISTRICT")
                city = pick("City", "Town", "CITY")
                country = pick("Country") or "India"
                email = pick("Email", "EMAIL")
                reg_code = pick("Lab Code", "Code", "Accession No")
                records.append({
                    "orgName": name,
                    "orgType": "Diagnostic Laboratory",
                    "orgCountry": country,
                    "orgState": state,
                    "orgDistrict": district,
                    "orgCity": city,
                    "email": email,
                    "regCode": reg_code or "NABL-LAB",
                    "accreditations": ["NABL Accreditation"],
                    "status": "approved",
                    "selectedMetrics": [],
                })
            return records
        except Exception:
            pass
        # Fallback to openpyxl
        try:
            from openpyxl import load_workbook  # type: ignore
            wb = load_workbook(filename=str(xlsx_path), read_only=True)
            ws = wb.active
            headers = [str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
            def get(row, *names):
                for n in names:
                    if n in headers:
                        idx = headers.index(n)
                        val = row[idx].value
                        return str(val) if val is not None else ""
                return ""
            records = []
            for row in ws.iter_rows(min_row=2):
                name = get(row, "Laboratory Name", "Lab Name", "Organization", "Name")
                state = get(row, "State", "STATE")
                district = get(row, "District", "DISTRICT")
                city = get(row, "City", "Town", "CITY")
                country = get(row, "Country") or "India"
                email = get(row, "Email", "EMAIL")
                reg_code = get(row, "Lab Code", "Code", "Accession No")
                records.append({
                    "orgName": name,
                    "orgType": "Diagnostic Laboratory",
                    "orgCountry": country,
                    "orgState": state,
                    "orgDistrict": district,
                    "orgCity": city,
                    "email": email,
                    "regCode": reg_code or "NABL-LAB",
                    "accreditations": ["NABL Accreditation"],
                    "status": "approved",
                    "selectedMetrics": [],
                })
            return records
        except Exception:
            return []
    except Exception:
        return []
APP_VERSION = "3"
